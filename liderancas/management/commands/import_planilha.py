"""
Importa a planilha CONTATOS LIDERANÇAS.xlsx para o CRM.

Abas → destinos:
- CTT LIDERANÇAS: coluna COORDENADOR → CoordenadorRegional;
  STATUS "LIDERANÇA" → CaboEleitoral; demais → Apoiador.
- LIDERANÇAS CHAPECÓ → CaboEleitoral (cidade Chapecó, bairro nas observações).
- EXECUTIVA → Apoiador político (cargo presidente_diretorio) + Cidade.presidente_diretorio.
- PREFEITOS → Apoiador político (cargo prefeito) + Cidade.prefeito_nome.
- VICE-PREFEITOS → Apoiador político (cargo vice_prefeito).
- VEREADORES → Apoiador político (cargo vereador).
- Uceff Lideranças / APOIADORES UCEFF → Apoiador comunitário.

Idempotente: usa get_or_create por chave normalizada (nome + cidade) e
apenas completa campos vazios/mescla observações em registros existentes.

Uso:
    python manage.py import_planilha "Planilha/CONTATOS LIDERANÇAS.xlsx" --dry-run
    python manage.py import_planilha "Planilha/CONTATOS LIDERANÇAS.xlsx"
"""
import re
import unicodedata
from collections import Counter
from difflib import get_close_matches

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from liderancas.models import (
    Cidade, Lideranca,
)

ORIGEM = 'Planilha CONTATOS LIDERANÇAS'

# Correções manuais de nomes de cidade encontrados na planilha
CITY_FIXES = {
    'compo ere': 'Campo Erê',
    'curutiba': 'Curitibanos',
    'entre rio': 'Entre Rios',
    'guariciaba': 'Guaraciaba',
    'faxinal do guedes': 'Faxinal dos Guedes',
    'dionisio': 'Dionísio Cerqueira',
    'herval do oeste': "Herval d'Oeste",
    'tajai': 'Itajaí',
    'sao miguel do oeste sc': 'São Miguel do Oeste',
    'caxambu': 'Caxambu do Sul',
    'smo': 'São Miguel do Oeste',
    'caxambu e planalto alegre': 'Caxambu do Sul',
    'santa ter. progresso': 'Santa Terezinha do Progresso',
    'sao lourenco': 'São Lourenço do Oeste',
    'vereador de vargeao': 'Vargeão',
    'riquesa': 'Riqueza',
    '- vereador pl princesa': 'Princesa',
}


def strip_accents(s):
    s = unicodedata.normalize('NFD', s)
    return ''.join(c for c in s if unicodedata.category(c) != 'Mn')


def norm(s):
    """Normaliza texto para comparação: sem acentos, minúsculo, espaços únicos."""
    return re.sub(r'\s+', ' ', strip_accents(str(s or '')).strip().lower())


def cell(row, idx):
    """Valor de célula como string limpa ('' para None/NULL)."""
    if idx >= len(row) or row[idx] is None:
        return ''
    v = str(row[idx]).strip()
    if v.upper() == 'NULL':
        return ''
    # remove artefato ".0" de números lidos como float
    if re.fullmatch(r'\d+\.0', v):
        v = v[:-2]
    return v


def clean_phone(raw):
    digits = re.sub(r'\D', '', cell_str(raw))
    # remove código do país
    if len(digits) >= 12 and digits.startswith('55'):
        digits = digits[2:]
    # 8-9 dígitos: assume DDD 49 (oeste de SC, origem da planilha)
    if 8 <= len(digits) <= 9:
        digits = '49' + digits
    if len(digits) < 10 or len(digits) > 11:
        return digits[:20] if digits else ''
    ddd, num = digits[:2], digits[2:]
    if len(num) == 9:
        return f'({ddd}) {num[:5]}-{num[5:]}'
    return f'({ddd}) {num[:4]}-{num[4:]}'


def cell_str(v):
    if v is None:
        return ''
    v = str(v).strip()
    if re.fullmatch(r'\d+\.0', v):
        v = v[:-2]
    return v


def clean_email(raw):
    v = cell_str(raw)
    if not v or v.upper() == 'NULL' or '@' not in v:
        return ''
    return v[:254]


def obs_block(pairs):
    """Monta texto de observações 'Chave: valor' apenas com valores presentes."""
    return '\n'.join(f'{k}: {v}' for k, v in pairs if v)


def merge_obs(existing, new_block):
    """Acrescenta linhas novas às observações sem duplicar."""
    if not new_block:
        return existing
    lines = [l for l in new_block.split('\n') if l and l not in (existing or '')]
    if not lines:
        return existing
    return (existing + '\n' if existing else '') + '\n'.join(lines)


class Command(BaseCommand):
    help = 'Importa a planilha CONTATOS LIDERANÇAS.xlsx (coordenadores, cabos, apoiadores).'

    def add_arguments(self, parser):
        parser.add_argument('arquivo', nargs='?', default='Planilha/CONTATOS LIDERANÇAS.xlsx')
        parser.add_argument('--dry-run', action='store_true', help='Simula a importação sem gravar nada.')

    # ---------- infra ----------

    def setup_cidades(self):
        self.cidades = {}
        for c in Cidade.objects.select_related('regiao'):
            self.cidades[norm(c.nome)] = c
        self.cidade_keys = list(self.cidades.keys())
        self.chapeco = self.cidades.get('chapeco')

    def find_cidade(self, nome):
        """Resolve nome de cidade para objeto Cidade (exato, correção manual ou fuzzy)."""
        key = norm(nome)
        if not key:
            return None
        key = re.sub(r'^l\s+', '', key)  # lixo "L ARVOREDO" da aba VEREADORES
        if key in self.cidades:
            return self.cidades[key]
        if key in CITY_FIXES:
            return self.cidades.get(norm(CITY_FIXES[key]))
        match = get_close_matches(key, self.cidade_keys, n=1, cutoff=0.86)
        if match:
            return self.cidades[match[0]]
        return None

    def pend(self, aba, linha, nome, motivo):
        self.pendencias.append((aba, linha, nome, motivo))

    def stat(self, key):
        self.stats[key] += 1

    # ---------- upserts ----------

    def upsert_apoiador(self, nome, cidade, tipo, defaults):
        key = ('apoiador', norm(nome), cidade.pk)
        if key in self.seen:
            obj = self.seen[key]
            self.fill(obj, defaults)
            return obj
        obj = Lideranca.all_objects.filter(
            papel='apoiador', cidade=cidade, nome__iexact=nome).first()
        if obj:
            self.fill(obj, defaults)
            self.stat('apoiadores atualizados')
        else:
            obj = Lideranca(papel='apoiador', nome=nome, cidade=cidade, tipo=tipo, **{
                k: v for k, v in defaults.items() if v
            })
            if not self.dry_run:
                obj.save()
            self.stat('apoiadores criados')
        self.seen[key] = obj
        return obj

    def upsert_cabo(self, nome, cidade, coordenador, defaults):
        key = ('cabo', norm(nome), cidade.pk)
        if key in self.seen:
            obj = self.seen[key]
            self.fill(obj, defaults)
            return obj
        obj = Lideranca.all_objects.filter(
            papel='cabo', cidade=cidade, nome__iexact=nome).first()
        if obj:
            self.fill(obj, defaults)
            self.stat('cabos atualizados')
        else:
            obj = Lideranca(papel='cabo', nome=nome, cidade=cidade,
                            coordenador_responsavel=coordenador, **{
                k: v for k, v in defaults.items() if v
            })
            if not self.dry_run:
                obj.save()
            self.stat('cabos criados')
        self.seen[key] = obj
        return obj

    def fill(self, obj, defaults):
        """Completa campos vazios e mescla observações; não sobrescreve dados existentes."""
        changed = []
        for field, value in defaults.items():
            if not value:
                continue
            if field == 'observacoes':
                merged = merge_obs(obj.observacoes, value)
                if merged != obj.observacoes:
                    obj.observacoes = merged
                    changed.append(field)
            elif not getattr(obj, field, None):
                setattr(obj, field, value)
                changed.append(field)
        if changed and not self.dry_run:
            obj.save(update_fields=changed if obj.pk else None)

    # ---------- abas ----------

    def import_ctt(self, ws):
        aba = 'CTT LIDERANÇAS'
        rows = list(ws.iter_rows(values_only=True))[1:]
        # 1ª passada: coordenadores e suas cidades
        coord_cidades = {}
        candidato_norm = norm(settings.CAMPANHA['CANDIDATO_PRIMEIRO_NOME'])
        for row in rows:
            coord = cell(row, 5)
            if not coord or norm(coord) == candidato_norm:
                continue
            cidade = self.find_cidade(cell(row, 0))
            coord_cidades.setdefault(coord, []).append(cidade)

        self.coordenadores = {}
        coords_existentes = {
            norm(c.nome): c
            for c in Lideranca.all_objects.filter(papel='coordenador')
        }
        for coord_nome, cidades in coord_cidades.items():
            validas = [c for c in cidades if c]
            base = Counter(c.pk for c in validas).most_common(1)
            cidade_base = next((c for c in validas if base and c.pk == base[0][0]), None) or self.chapeco
            if not cidade_base:
                self.pend('CTT LIDERANÇAS', 0, coord_nome, 'coordenador sem cidade-base resolvível')
                continue
            obj = coords_existentes.get(norm(coord_nome))
            if not obj:
                obj = Lideranca(
                    papel='coordenador',
                    nome=coord_nome.title(),
                    regiao=cidade_base.regiao,
                    cidade=cidade_base,
                    observacoes=f'Origem: {ORIGEM} (coluna COORDENADOR)',
                )
                if not self.dry_run:
                    obj.save()
                self.stat('coordenadores criados')
            self.coordenadores[norm(coord_nome)] = obj

        # 2ª passada: cabos e apoiadores
        for i, row in enumerate(rows, start=2):
            nome = cell(row, 1)
            if not nome:
                continue
            cidade = self.find_cidade(cell(row, 0))
            if not cidade:
                self.pend(aba, i, nome, f'cidade não identificada: "{cell(row, 0)}"')
                continue
            coord_raw = cell(row, 5)
            coordenador = self.coordenadores.get(norm(coord_raw)) if coord_raw else None
            origem = settings.CAMPANHA['CANDIDATO_PRIMEIRO_NOME'] if norm(coord_raw) == norm(settings.CAMPANHA['CANDIDATO_PRIMEIRO_NOME']) else ORIGEM
            obs = obs_block([
                ('Função', cell(row, 3)),
                ('Status planilha', cell(row, 4)),
                ('Fechado', cell(row, 6)),
                ('Visitas', cell(row, 8)),
                ('Obs', cell(row, 9)),
                ('Veio por', cell(row, 10)),
                ('Origem', ORIGEM),
            ])
            defaults = {
                'telefone': clean_phone(row[2] if len(row) > 2 else ''),
                'instagram': cell(row, 7)[:100],
                'observacoes': obs,
            }
            status = norm(cell(row, 4))
            if 'lideranca' in status:
                defaults['email'] = ''
                self.upsert_cabo(nome, cidade, coordenador, defaults)
            else:
                funcao = norm(cell(row, 3))
                tipo, cargo = 'comunitario', ''
                if 'verea' in funcao:
                    tipo, cargo = 'politico', 'ex_politico' if 'ex' in funcao else 'vereador'
                elif 'vice' in funcao and 'prefeit' in funcao:
                    tipo, cargo = 'politico', 'vice_prefeito'
                elif 'prefeit' in funcao:
                    tipo, cargo = 'politico', 'ex_politico' if 'ex' in funcao else 'prefeito'
                elif 'empres' in funcao:
                    tipo = 'empresarial'
                defaults.update({'cargo': cargo, 'origem_contato': origem})
                self.upsert_apoiador(nome, cidade, tipo, defaults)

    def import_liderancas_chapeco(self, ws):
        aba = 'LIDERANÇAS CHAPECÓ'
        if not self.chapeco:
            self.pend(aba, 0, '-', 'cidade Chapecó não encontrada no banco')
            return
        for i, row in enumerate(list(ws.iter_rows(values_only=True))[1:], start=2):
            nome = cell(row, 1)
            if not nome:
                continue
            obs = obs_block([
                ('Bairro', cell(row, 0)),
                ('Função', cell(row, 3)),
                ('Origem', f'{ORIGEM} (aba LIDERANÇAS CHAPECÓ)'),
            ])
            self.upsert_cabo(nome, self.chapeco, None, {
                'telefone': clean_phone(row[2] if len(row) > 2 else ''),
                'observacoes': obs,
            })

    def import_politicos(self, ws, aba, header_rows, col_cidade, col_nome, col_tel,
                         cargo, col_regional=None, col_email=None, update_field=None):
        for i, row in enumerate(list(ws.iter_rows(values_only=True))[header_rows:], start=header_rows + 1):
            nome = cell(row, col_nome)
            if not nome:
                continue
            nome = re.sub(r'\s*-\s*$', '', nome).strip()
            cidade_raw = cell(row, col_cidade)
            if aba == 'EXECUTIVA':
                cidade_raw = cidade_raw.split('/')[0].strip()
            cidade = self.find_cidade(cidade_raw)
            if not cidade:
                self.pend(aba, i, nome, f'cidade não identificada: "{cidade_raw}"')
                continue
            obs = obs_block([
                ('Regional', cell(row, col_regional) if col_regional is not None else ''),
                ('Origem', f'{ORIGEM} (aba {aba})'),
            ])
            self.upsert_apoiador(nome, cidade, 'politico', {
                'cargo': cargo,
                'telefone': clean_phone(row[col_tel] if len(row) > col_tel else ''),
                'email': clean_email(row[col_email]) if col_email is not None and len(row) > col_email else '',
                'origem_contato': ORIGEM,
                'grau_influencia': 'alto',
                'observacoes': obs,
            })
            if update_field and not self.dry_run:
                if not getattr(cidade, update_field, None):
                    setattr(cidade, update_field, nome[:200])
                    cidade.save(update_fields=[update_field])

    def import_uceff(self, ws, aba, header_rows):
        for i, row in enumerate(list(ws.iter_rows(values_only=True))[header_rows:], start=header_rows + 1):
            nome = cell(row, 0)
            if not nome:
                continue
            unidade = cell(row, 1)
            cidade = self.find_cidade(unidade) or self.chapeco
            if not cidade:
                self.pend(aba, i, nome, f'cidade não identificada: "{unidade}"')
                continue
            obs = obs_block([
                ('Unidade UCEFF', unidade or 'não informada'),
                ('Função/Curso', cell(row, 3)),
                ('Origem', f'{ORIGEM} (aba {aba})'),
            ])
            self.upsert_apoiador(nome, cidade, 'comunitario', {
                'telefone': clean_phone(row[5] if len(row) > 5 else ''),
                'instagram': cell(row, 4)[:100],
                'origem_contato': f'UCEFF ({aba})',
                'observacoes': obs,
            })

    # ---------- main ----------

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            raise CommandError('openpyxl não instalado. Rode: pip install openpyxl')

        self.dry_run = options['dry_run']
        self.stats = Counter()
        self.pendencias = []
        self.seen = {}

        try:
            wb = openpyxl.load_workbook(options['arquivo'], read_only=True, data_only=True)
        except FileNotFoundError:
            raise CommandError(f'Arquivo não encontrado: {options["arquivo"]}')

        self.setup_cidades()
        if not self.cidades:
            raise CommandError('Nenhuma cidade no banco. Rode antes: python manage.py load_sc')

        if self.dry_run:
            self.stdout.write(self.style.WARNING('== DRY-RUN: nada será gravado =='))

        with transaction.atomic():
            self.import_ctt(wb['CTT LIDERANÇAS'])
            self.import_liderancas_chapeco(wb['LIDERANÇAS CHAPECÓ'])
            self.import_politicos(wb['EXECUTIVA'], 'EXECUTIVA', 1, 0, 1, 3,
                                  'presidente_diretorio', col_email=4, update_field='presidente_diretorio')
            self.import_politicos(wb['PREFEITOS'], 'PREFEITOS', 2, 1, 2, 3,
                                  'prefeito', col_regional=0, update_field='prefeito_nome')
            self.import_politicos(wb['VICE-PREFEITOS'], 'VICE-PREFEITOS', 2, 1, 2, 3,
                                  'vice_prefeito', col_regional=0)
            self.import_politicos(wb['VEREADORES'], 'VEREADORES', 2, 1, 2, 3,
                                  'vereador', col_regional=0)
            self.import_uceff(wb['Uceff Lideranças'], 'Uceff Lideranças', 2)
            self.import_uceff(wb['APOIADORES UCEFF'], 'APOIADORES UCEFF', 2)

            if self.dry_run:
                transaction.set_rollback(True)

        self.stdout.write('')
        self.stdout.write(self.style.MIGRATE_HEADING('== Resultado =='))
        for key in sorted(self.stats):
            self.stdout.write(f'  {key}: {self.stats[key]}')

        if self.pendencias:
            self.stdout.write('')
            self.stdout.write(self.style.WARNING(
                f'== {len(self.pendencias)} pendências (linhas não importadas) =='
            ))
            for aba, linha, nome, motivo in self.pendencias:
                self.stdout.write(f'  [{aba} linha {linha}] {nome}: {motivo}')
            import csv as csv_mod
            import os
            report = os.path.join(os.path.dirname(options['arquivo']) or '.', 'import_pendencias.csv')
            with open(report, 'w', newline='', encoding='utf-8-sig') as fh:
                w = csv_mod.writer(fh)
                w.writerow(['Aba', 'Linha', 'Nome', 'Motivo'])
                w.writerows(self.pendencias)
            self.stdout.write(self.style.WARNING(f'Relatório salvo em: {report}'))

        if self.dry_run:
            self.stdout.write(self.style.WARNING('\nDRY-RUN concluído — nada foi gravado.'))
        else:
            self.stdout.write(self.style.SUCCESS('\nImportação concluída.'))
