"""
Exporta UMA planilha .xlsx unificando todos os cadastros de pessoas do CRM
numa única aba "Base Unificada", com discriminador "Tipo de Registro".

Incorpora: Coordenadores Regionais, Cabos Eleitorais, Apoiadores, Voluntários,
Egressos, Lassberg e Demandas do Eleitor (Promessas). NÃO inclui doadores.

Recursos da planilha:
- Cabeçalho congelado (freeze) + colunas ID/Tipo/Nome travadas na horizontal.
- AutoFiltro em todas as colunas.
- Dropdowns de validação (aba "Listas") para Tipo, Prioridade, Frequência,
  Status/Situação, Cargo, Categoria, Grau de Influência, Em Atraso?, Ativo.
- Destaque vermelho na linha quando "Em Atraso? = Sim".
- "Última Interação" / "Em Atraso?" replicam a lógica da Fila de Relacionamento
  (FREQ_PRAZOS: semanal=7, quinzenal=15, mensal=30, eventual=90 dias).

Uso:
    python manage.py exportar_base_unificada
    python manage.py exportar_base_unificada --output "media/base.xlsx"
    python manage.py exportar_base_unificada --incluir-inativos
"""
from django.core.management.base import BaseCommand
from django.db.models import Max
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

from liderancas.models import (
    Lideranca, Voluntario, Egresso, Lassberg,
)
from tarefas.models import Promessa

# Prazos de relacionamento — espelham liderancas/views.py:FREQ_PRAZOS
FREQ_PRAZOS = {'semanal': 7, 'quinzenal': 15, 'mensal': 30, 'eventual': 90}

# Ordem das colunas da Base Unificada (38 colunas)
HEADERS = [
    'ID', 'Tipo de Registro', 'Nome',
    'Telefone', 'E-mail', 'Instagram / Redes',
    'Macrorregião', 'Região', 'Cidade', 'UF', 'Bairro / Linha / Zona',
    'Coordenador Responsável', 'Cargo', 'Categoria do Apoiador',
    'Prioridade', 'Frequência de Contato', 'Grau de Influência', 'Status / Situação',
    'Última Interação', 'Dias Sem Contato', 'Em Atraso?', 'Próximo Contato Sugerido',
    'Votos de Referência', 'Meta de Votos a Transferir', 'Origem do Contato',
    'Curso', 'Instituição', 'Situação do Curso',
    'Demanda', 'Responsável pela Entrega', 'Data de Registro', 'Data de Entrega',
    'Cadastrado por', 'Atualizado por', 'Data de Cadastro', 'Última Atualização',
    'Ativo', 'Observações',
]

# Larguras de coluna (índice 0-based) → largura
COL_WIDTHS = {
    0: 10, 1: 18, 2: 28, 3: 16, 4: 26, 5: 18, 6: 16, 7: 18, 8: 18, 9: 6,
    10: 22, 11: 24, 12: 18, 13: 22, 14: 11, 15: 16, 16: 16, 17: 15, 18: 14,
    19: 14, 20: 11, 21: 18, 22: 14, 23: 16, 24: 22, 25: 18, 26: 20, 27: 18,
    28: 36, 29: 22, 30: 14, 31: 14, 32: 18, 33: 18, 34: 14, 35: 16, 36: 8, 37: 40,
}

# Listas de validação (dropdowns) — (cabeçalho, valores)
LISTAS = [
    ('Tipo de Registro', ['Coordenador Regional', 'Cabo Eleitoral', 'Apoiador',
                          'Voluntário', 'Egresso', 'Lassberg', 'Demanda']),
    ('Prioridade', ['Alta', 'Média', 'Baixa']),
    ('Frequência', ['Semanal', 'Quinzenal', 'Mensal', 'Eventual']),
    ('Status / Situação', ['Ativo', 'Inativo', 'Pendente',
                           'Registrada', 'Em andamento', 'Entregue', 'Cancelada']),
    ('Cargo', ['Prefeito', 'Vice-Prefeito', 'Vereador',
               'Presidente Diretório', 'Ex-Político', 'Outro']),
    ('Categoria do Apoiador', ['Apoiador Político', 'Apoiador Empresarial',
                               'Apoiador Comunitário', 'Líder de Associação',
                               'Eleitor Estratégico', 'Imprensa', 'Apoiador PWA']),
    ('Grau de Influência', ['Alto', 'Médio', 'Baixo']),
    ('Em Atraso?', ['Sim', 'Não']),
    ('Ativo', ['Sim', 'Não']),
]

# Mapeia coluna da Base Unificada (1-based) → cabeçalho da lista de validação
VALIDACAO_COLS = {
    'Tipo de Registro': 'Tipo de Registro',
    'Prioridade': 'Prioridade',
    'Frequência de Contato': 'Frequência',
    'Status / Situação': 'Status / Situação',
    'Cargo': 'Cargo',
    'Categoria do Apoiador': 'Categoria do Apoiador',
    'Grau de Influência': 'Grau de Influência',
    'Em Atraso?': 'Em Atraso?',
    'Ativo': 'Ativo',
}


class Command(BaseCommand):
    help = 'Exporta a Base Unificada (.xlsx) com todos os cadastros de pessoas e demandas.'

    def add_arguments(self, parser):
        parser.add_argument('--output', default=None, help='Caminho do .xlsx de saída.')
        parser.add_argument('--incluir-inativos', action='store_true',
                            help='Inclui também registros soft-deleted (Ativo = Não).')
        parser.add_argument('--incluir-pendentes', action='store_true',
                            help='Inclui leads do app ainda não aprovados (default: só aprovados).')

    def handle(self, *args, **options):
        self.agora = timezone.now()
        self.hoje = timezone.localdate()
        incluir_inativos = options['incluir_inativos']
        # Planilha oficial = só aprovados; rejeitados nunca entram.
        self.aprovacoes = ['aprovado', 'pendente'] if options['incluir_pendentes'] else ['aprovado']
        output = options['output'] or f'base_unificada_{self.hoje:%Y%m%d}.xlsx'

        rows = []
        rows += self._coordenadores(incluir_inativos)
        rows += self._cabos(incluir_inativos)
        rows += self._apoiadores(incluir_inativos)
        rows += self._voluntarios()
        rows += self._egressos(incluir_inativos)
        rows += self._lassberg(incluir_inativos)
        rows += self._demandas()

        self._write_workbook(rows, output)

        self.stdout.write(self.style.SUCCESS(
            f'✓ {len(rows)} registros exportados para "{output}".'
        ))

    # ---------- helpers de geografia / relacionamento ----------

    def _geo(self, cidade, regiao=None):
        """Retorna (macrorregião, região, cidade_nome) a partir de uma Cidade
        (ou de uma Região avulsa quando não há cidade)."""
        if cidade is not None:
            regiao = cidade.regiao
            cidade_nome = cidade.nome
        else:
            cidade_nome = ''
        if regiao is not None:
            macro = regiao.macro_regiao.nome if regiao.macro_regiao_id else ''
            regiao_nome = regiao.nome
        else:
            macro = regiao_nome = ''
        return macro, regiao_nome, cidade_nome

    def _coord_inferido(self, regiao):
        """Coordenador inferido pela região (para tipos sem FK direta)."""
        if regiao is None:
            return ''
        coord = Lideranca.objects.filter(papel='coordenador', regiao=regiao).first()
        return f'{coord.nome} (inferido)' if coord else ''

    def _relacionamento(self, ultima, frequencia):
        """Replica a Fila: (última_data, dias_sem_contato, em_atraso, próximo_sugerido)."""
        prazo = FREQ_PRAZOS.get(frequencia, 30)
        if ultima is None:
            return None, 'Nunca', 'Sim', None
        dias = (self.agora - ultima).days
        em_atraso = 'Sim' if dias > prazo else 'Não'
        proximo = timezone.localtime(ultima).date() + timezone.timedelta(days=prazo)
        return timezone.localtime(ultima).date(), dias, em_atraso, proximo

    @staticmethod
    def _user(u):
        if not u:
            return ''
        return u.get_full_name() or u.get_username()

    @staticmethod
    def _ativo(obj):
        return 'Não' if getattr(obj, 'is_active', True) is False else 'Sim'

    def _base_row(self):
        return {h: '' for h in HEADERS}

    # ---------- construtores por tipo ----------

    def _coordenadores(self, incluir_inativos):
        qs = (Lideranca.all_objects if incluir_inativos
              else Lideranca.objects).filter(papel='coordenador', aprovacao__in=self.aprovacoes)
        qs = qs.select_related('regiao', 'regiao__macro_regiao', 'cidade',
                               'cidade__regiao__macro_regiao',
                               'cadastrado_por', 'atualizado_por')
        qs = qs.annotate(ultima=Max('interacoes__data'))
        out = []
        for c in qs:
            macro, regiao, cidade = self._geo(c.cidade)
            ult, dias, atraso, prox = self._relacionamento(c.ultima, c.frequencia_relacionamento)
            r = self._base_row()
            r.update({
                'ID': f'COORD-{c.pk}', 'Tipo de Registro': 'Coordenador Regional',
                'Nome': c.nome, 'Telefone': c.telefone, 'E-mail': c.email,
                'Instagram / Redes': c.instagram,
                'Macrorregião': macro, 'Região': regiao, 'Cidade': cidade, 'UF': 'SC',
                'Prioridade': c.get_prioridade_display(),
                'Frequência de Contato': c.get_frequencia_relacionamento_display(),
                'Última Interação': ult, 'Dias Sem Contato': dias,
                'Em Atraso?': atraso, 'Próximo Contato Sugerido': prox,
                'Cadastrado por': self._user(c.cadastrado_por),
                'Atualizado por': self._user(c.atualizado_por),
                'Data de Cadastro': c.created_at.date() if c.created_at else '',
                'Última Atualização': c.updated_at.date() if c.updated_at else '',
                'Ativo': self._ativo(c), 'Observações': c.observacoes,
            })
            out.append(r)
        return out

    def _cabos(self, incluir_inativos):
        qs = (Lideranca.all_objects if incluir_inativos
              else Lideranca.objects).filter(papel='cabo', aprovacao__in=self.aprovacoes)
        qs = qs.select_related('cidade', 'cidade__regiao__macro_regiao',
                               'coordenador_responsavel',
                               'cadastrado_por', 'atualizado_por')
        qs = qs.annotate(ultima=Max('interacoes__data'))
        out = []
        for c in qs:
            macro, regiao, cidade = self._geo(c.cidade)
            ult, dias, atraso, prox = self._relacionamento(c.ultima, c.frequencia_relacionamento)
            r = self._base_row()
            r.update({
                'ID': f'CABO-{c.pk}', 'Tipo de Registro': 'Cabo Eleitoral',
                'Nome': c.nome, 'Telefone': c.telefone, 'E-mail': c.email,
                'Instagram / Redes': c.instagram,
                'Macrorregião': macro, 'Região': regiao, 'Cidade': cidade, 'UF': 'SC',
                'Coordenador Responsável': c.coordenador_responsavel.nome if c.coordenador_responsavel_id else '',
                'Prioridade': c.get_prioridade_display(),
                'Frequência de Contato': c.get_frequencia_relacionamento_display(),
                'Última Interação': ult, 'Dias Sem Contato': dias,
                'Em Atraso?': atraso, 'Próximo Contato Sugerido': prox,
                'Cadastrado por': self._user(c.cadastrado_por),
                'Atualizado por': self._user(c.atualizado_por),
                'Data de Cadastro': c.created_at.date() if c.created_at else '',
                'Última Atualização': c.updated_at.date() if c.updated_at else '',
                'Ativo': self._ativo(c), 'Observações': c.observacoes,
            })
            out.append(r)
        return out

    def _apoiadores(self, incluir_inativos):
        qs = (Lideranca.all_objects if incluir_inativos
              else Lideranca.objects).filter(papel='apoiador', aprovacao__in=self.aprovacoes)
        qs = qs.select_related('cidade', 'cidade__regiao__macro_regiao',
                               'cadastrado_por', 'atualizado_por')
        qs = qs.annotate(ultima=Max('interacoes__data'))
        out = []
        for a in qs:
            macro, regiao, cidade = self._geo(a.cidade)
            ult, dias, atraso, prox = self._relacionamento(a.ultima, a.frequencia_relacionamento)
            r = self._base_row()
            r.update({
                'ID': f'APO-{a.pk}', 'Tipo de Registro': 'Apoiador',
                'Nome': a.nome, 'Telefone': a.telefone, 'E-mail': a.email,
                'Instagram / Redes': a.instagram,
                'Macrorregião': macro, 'Região': regiao, 'Cidade': cidade, 'UF': 'SC',
                'Coordenador Responsável': self._coord_inferido(a.cidade.regiao if a.cidade_id else None),
                'Cargo': a.get_cargo_display() if a.cargo else '',
                'Categoria do Apoiador': a.get_tipo_display(),
                'Prioridade': a.get_prioridade_display(),
                'Frequência de Contato': a.get_frequencia_relacionamento_display(),
                'Grau de Influência': a.get_grau_influencia_display(),
                'Status / Situação': a.get_status_display(),
                'Última Interação': ult, 'Dias Sem Contato': dias,
                'Em Atraso?': atraso, 'Próximo Contato Sugerido': prox,
                'Votos de Referência': a.votos_referencia,
                'Meta de Votos a Transferir': a.meta_votos_transferir,
                'Origem do Contato': a.origem_contato,
                'Cadastrado por': self._user(a.cadastrado_por),
                'Atualizado por': self._user(a.atualizado_por),
                'Data de Cadastro': a.created_at.date() if a.created_at else '',
                'Última Atualização': a.updated_at.date() if a.updated_at else '',
                'Ativo': self._ativo(a), 'Observações': a.observacoes,
            })
            out.append(r)
        return out

    def _voluntarios(self):
        qs = Voluntario.objects.select_related(
            'cidade', 'cidade__regiao__macro_regiao', 'regiao__macro_regiao', 'cadastrado_por')
        out = []
        for v in qs:
            macro, regiao, cidade = self._geo(v.cidade, v.regiao)
            disp = ', '.join(v.get_disponibilidades_display())
            obs = v.observacoes
            if disp:
                obs = f'Disponibilidades: {disp}' + (f' | {obs}' if obs else '')
            r = self._base_row()
            r.update({
                'ID': f'VOL-{v.pk}', 'Tipo de Registro': 'Voluntário',
                'Nome': v.nome, 'Telefone': v.telefone,
                'Macrorregião': macro, 'Região': regiao, 'Cidade': cidade, 'UF': 'SC',
                'Bairro / Linha / Zona': v.endereco,
                'Coordenador Responsável': self._coord_inferido(
                    v.cidade.regiao if v.cidade_id else v.regiao),
                'Cadastrado por': self._user(v.cadastrado_por),
                'Data de Cadastro': v.created_at.date() if v.created_at else '',
                'Ativo': 'Sim', 'Observações': obs,
            })
            out.append(r)
        return out

    def _egressos(self, incluir_inativos):
        qs = (Egresso.all_objects if incluir_inativos else Egresso.objects)
        qs = qs.select_related('cidade', 'cidade__regiao__macro_regiao',
                               'cadastrado_por', 'atualizado_por')
        qs = qs.annotate(ultima=Max('interacoes__data'))
        out = []
        for e in qs:
            if e.cidade_id:
                macro, regiao, cidade = self._geo(e.cidade)
                uf = 'SC'
            else:
                macro = regiao = ''
                cidade = e.cidade_nome
                uf = e.estado
            ult = timezone.localtime(e.ultima).date() if e.ultima else None
            dias = (self.agora - e.ultima).days if e.ultima else ''
            r = self._base_row()
            r.update({
                'ID': f'EGR-{e.pk}', 'Tipo de Registro': 'Egresso',
                'Nome': e.nome, 'Telefone': e.telefone, 'E-mail': e.email,
                'Instagram / Redes': e.instagram,
                'Macrorregião': macro, 'Região': regiao, 'Cidade': cidade, 'UF': uf,
                'Última Interação': ult, 'Dias Sem Contato': dias,
                'Curso': e.curso, 'Instituição': e.instituicao,
                'Situação do Curso': e.situacao_curso,
                'Cadastrado por': self._user(e.cadastrado_por),
                'Atualizado por': self._user(e.atualizado_por),
                'Data de Cadastro': e.created_at.date() if e.created_at else '',
                'Última Atualização': e.updated_at.date() if e.updated_at else '',
                'Ativo': self._ativo(e), 'Observações': e.observacoes,
            })
            out.append(r)
        return out

    def _lassberg(self, incluir_inativos):
        qs = (Lassberg.all_objects if incluir_inativos else Lassberg.objects)
        qs = qs.select_related('cidade', 'cidade__regiao__macro_regiao',
                               'cadastrado_por', 'atualizado_por')
        qs = qs.annotate(ultima=Max('interacoes__data'))
        out = []
        for l in qs:
            if l.cidade_id:
                macro, regiao, cidade = self._geo(l.cidade)
                uf = 'SC'
            else:
                macro = regiao = ''
                cidade = l.cidade_nome
                uf = l.estado
            ult = timezone.localtime(l.ultima).date() if l.ultima else None
            dias = (self.agora - l.ultima).days if l.ultima else ''
            r = self._base_row()
            r.update({
                'ID': f'LAS-{l.pk}', 'Tipo de Registro': 'Lassberg',
                'Nome': l.nome, 'Telefone': l.telefone, 'E-mail': l.email,
                'Macrorregião': macro, 'Região': regiao, 'Cidade': cidade, 'UF': uf,
                'Última Interação': ult, 'Dias Sem Contato': dias,
                'Cadastrado por': self._user(l.cadastrado_por),
                'Atualizado por': self._user(l.atualizado_por),
                'Data de Cadastro': l.created_at.date() if l.created_at else '',
                'Última Atualização': l.updated_at.date() if l.updated_at else '',
                'Ativo': self._ativo(l), 'Observações': l.observacoes,
            })
            out.append(r)
        return out

    def _demandas(self):
        qs = Promessa.objects.select_related(
            'cidade', 'cidade__regiao__macro_regiao', 'cadastrado_por')
        out = []
        for p in qs:
            macro, regiao, cidade = self._geo(p.cidade)
            r = self._base_row()
            r.update({
                'ID': f'DEM-{p.pk}', 'Tipo de Registro': 'Demanda',
                'Nome': p.solicitante,
                'Macrorregião': macro, 'Região': regiao, 'Cidade': cidade, 'UF': 'SC',
                'Bairro / Linha / Zona': p.bairro_linha,
                'Status / Situação': p.get_status_display(),
                'Demanda': p.descricao,
                'Responsável pela Entrega': p.responsavel,
                'Data de Registro': p.data_registro,
                'Data de Entrega': p.data_entrega or '',
                'Cadastrado por': self._user(p.cadastrado_por),
                'Data de Cadastro': p.created_at.date() if p.created_at else '',
                'Última Atualização': p.updated_at.date() if p.updated_at else '',
                'Ativo': 'Sim', 'Observações': p.observacoes,
            })
            out.append(r)
        return out

    # ---------- escrita do workbook ----------

    def _write_workbook(self, rows, output):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Base Unificada'

        header_fill = PatternFill('solid', fgColor='1F3A2E')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        thin = Side(style='thin', color='D9D9D9')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        date_cols = {'Última Interação', 'Próximo Contato Sugerido',
                     'Data de Registro', 'Data de Entrega',
                     'Data de Cadastro', 'Última Atualização'}

        # cabeçalho
        for col, head in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=head)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        ws.row_dimensions[1].height = 30

        # dados
        for i, row in enumerate(rows, start=2):
            for col, head in enumerate(HEADERS, start=1):
                val = row.get(head, '')
                cell = ws.cell(row=i, column=col, value=val if val != '' else None)
                cell.border = border
                cell.alignment = Alignment(vertical='center',
                                           wrap_text=(head == 'Observações' or head == 'Demanda'))
                if head in date_cols and val:
                    cell.number_format = 'DD/MM/YYYY'

        n_rows = len(rows)
        last_col = get_column_letter(len(HEADERS))
        last_row = max(n_rows + 1, 2)

        # larguras
        for idx, width in COL_WIDTHS.items():
            ws.column_dimensions[get_column_letter(idx + 1)].width = width

        # freeze: trava cabeçalho (linha 1) + colunas ID/Tipo/Nome (A-C)
        ws.freeze_panes = 'D2'

        # autofiltro
        ws.auto_filter.ref = f'A1:{last_col}{last_row}'

        # aba de listas para dropdowns
        ws_listas = wb.create_sheet('Listas')
        col_listas = {}
        for j, (titulo, valores) in enumerate(LISTAS, start=1):
            letra = get_column_letter(j)
            ws_listas.cell(row=1, column=j, value=titulo).font = Font(bold=True)
            for k, v in enumerate(valores, start=2):
                ws_listas.cell(row=k, column=j, value=v)
            col_listas[titulo] = f"Listas!${letra}$2:${letra}${len(valores) + 1}"
            ws_listas.column_dimensions[letra].width = 22

        # validações (dropdowns) na Base Unificada
        if n_rows:
            for head, lista_titulo in VALIDACAO_COLS.items():
                col_idx = HEADERS.index(head) + 1
                letra = get_column_letter(col_idx)
                dv = DataValidation(type='list', formula1=col_listas[lista_titulo],
                                    allow_blank=True, showDropDown=False)
                dv.error = 'Selecione um valor da lista.'
                dv.errorTitle = 'Valor inválido'
                dv.prompt = f'Escolha: {lista_titulo}'
                ws.add_data_validation(dv)
                dv.add(f'{letra}2:{letra}{last_row}')

            # destaque vermelho da linha quando "Em Atraso? = Sim"
            atraso_col = get_column_letter(HEADERS.index('Em Atraso?') + 1)
            regra_fill = PatternFill('solid', fgColor='FCE4E4')
            regra_font = Font(color='B00020')
            ws.conditional_formatting.add(
                f'A2:{last_col}{last_row}',
                FormulaRule(formula=[f'${atraso_col}2="Sim"'],
                            fill=regra_fill, font=regra_font, stopIfTrue=False),
            )

        wb.save(output)
