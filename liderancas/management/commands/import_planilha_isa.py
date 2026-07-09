"""Importa a PLANILHA CENTRAL ISA para o CRM.

- BASE DA ESTRUTURA → lideranças APROVADAS (base rica, 25 colunas).
- Nomes que só existem na BASE DE CONTATOS VERDADEIRA → leads PENDENTES
  (aprovacao='pendente'), com o que dá para mapear + as diferenças jogadas em
  Observações com aviso "⚠️ REVISAR".

Regras (decididas com o responsável):
- Importa TODAS as UFs. Contato de fora de SC (ou sem cidade na planilha) entra
  SEM cidade (campo opcional); a cidade/UF crua vai para as Observações.
- Atendente casa por USERNAME (sem acento/maiúscula). Sem match → vazio + nota.
- "Nível de engajamento" (VERDADEIRA) NÃO vira `nivel` (valores inconsistentes) —
  vai só para Observações. "Rede social" (VERDADEIRA) → Canal.
- Idempotente: dedup por telefone; sem telefone, por nome+cidade/UF. Reimportar
  atualiza, não duplica.

Uso:
    python manage.py import_planilha_isa --arquivo "/caminho/PLANILHA.xlsx"          # DRY-RUN
    python manage.py import_planilha_isa --arquivo "..." --sim                        # grava
"""
import re
import unicodedata
from collections import Counter
from datetime import date, datetime

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from liderancas.models import Cidade, Lideranca, InteracaoLog
from usuarios.models import Usuario


# ── Normalização ─────────────────────────────────────────────────

def norm(s):
    s = ''.join(c for c in unicodedata.normalize('NFD', str(s or ''))
                if unicodedata.category(c) != 'Mn')
    return re.sub(r'\s+', ' ', s.strip().lower())


def cell_str(v):
    if v is None:
        return ''
    v = str(v).strip()
    if v.upper() == 'NULL':
        return ''
    if re.fullmatch(r'\d+\.0', v):
        v = v[:-2]
    return v


def clean_phone(raw):
    d = re.sub(r'\D', '', cell_str(raw))
    if len(d) >= 12 and d.startswith('55'):
        d = d[2:]
    if 8 <= len(d) <= 9:
        d = '49' + d
    if len(d) < 10 or len(d) > 11:
        return d[:20] if d else ''
    ddd, num = d[:2], d[2:]
    return f'({ddd}) {num[:5]}-{num[5:]}' if len(num) == 9 else f'({ddd}) {num[:4]}-{num[4:]}'


def clean_email(raw):
    v = cell_str(raw)
    return v[:254] if v and '@' in v else ''


def to_bool(raw):
    return norm(raw) in ('sim', 's', 'true', '1', 'x', 'ok')


def to_date(raw):
    if raw in (None, ''):
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    s = cell_str(raw)
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def to_idade(raw):
    m = re.search(r'\d{1,3}', cell_str(raw))
    if not m:
        return None
    v = int(m.group())
    return v if 0 < v < 130 else None


def map_voto(raw):
    n = norm(raw)
    return {'sim': 'sim', 'talvez': 'talvez', 'nao': 'nao',
            'nao contactado': 'nao_contactado'}.get(n, '')


def map_nivel(raw):
    n = norm(raw)
    if n.startswith('multiplicador') or n in ('lider', 'líder'):
        return 'multiplicador'
    if n.startswith('lead'):
        return 'lead'
    if n.startswith('eleitor'):
        return 'eleitor'
    if n.startswith('volunt'):
        return 'voluntario'
    return ''


def map_canal(raw):
    n = norm(raw)
    if 'whats' in n:
        return 'whatsapp'
    if 'insta' in n:
        return 'instagram'
    if 'face' in n:
        return 'facebook'
    if 'tiktok' in n or 'tik tok' in n:
        return 'tiktok'
    if 'twitter' in n or n == 'x':
        return 'twitter'
    return ''


# Título normalizado → chave interna, por aba
TITULOS_ESTRUTURA = {
    'atendente': 'atendente', 'voto': 'voto', 'nivel': 'nivel', 'nome': 'nome',
    'cidade': 'cidade', 'uf': 'uf', 'contato feito?': 'contato_feito',
    'data contato': 'data_contato', 'canal do ultimo atendimento': 'canal',
    'telefone': 'telefone', 'ja mandou link da vaquinha?': 'vaquinha', 'doou?': 'doou',
    'observacoes gerais': 'observacoes', 'filiado a algum partido?': 'filiado',
    'quem e o eleitor?': 'quem_eleitor', 'como chegou?': 'origem', 'instagram': 'instagram',
    'facebook': 'facebook', 'e-mail': 'email', 'endereco (pra entrega de material)': 'endereco',
    'material entregue?': 'material', 'idade': 'idade', 'segmentos/interesses': 'segmentos',
    'historico do atendimento': 'historico',
}
TITULOS_VERDADEIRA = {
    'voto': 'voto', 'nome': 'nome', 'cidade': 'cidade', 'uf': 'uf',
    'contato feito?': 'contato_feito', 'data contato': 'data_contato',
    'ja mandou link da vaquinha?': 'vaquinha', 'rede social': 'canal', 'telefone': 'telefone',
    'atendente': 'atendente', 'observacao': 'observacoes', 'quem e?': 'quem_eleitor',
    'como chegou?': 'origem', 'rede social 2': 'rede2', 'nivel engajamento': 'engajamento',
    'e-mail': 'email', 'endereco': 'endereco', 'idade': 'idade', 'segmento/interesses': 'segmentos',
}


class Command(BaseCommand):
    help = 'Importa a PLANILHA CENTRAL ISA (ESTRUTURA aprovada + 64 exclusivos como leads pendentes).'

    def add_arguments(self, parser):
        parser.add_argument('--arquivo', required=True)
        parser.add_argument('--sim', action='store_true',
                            help='Grava de fato. Sem isso, DRY-RUN.')

    def handle(self, *args, **opts):
        import openpyxl
        executar = opts['sim']
        self.stats = Counter()
        self.pend = []
        self.atendentes_sem = Counter()
        self.cidades_sem = Counter()

        wb = openpyxl.load_workbook(opts['arquivo'], read_only=True, data_only=True)
        for aba in ('BASE DA ESTRUTURA', 'BASE DE CONTATOS VERDADEIRA'):
            if aba not in wb.sheetnames:
                raise CommandError(f'Aba "{aba}" não existe.')

        # Índices auxiliares
        _cids = list(Cidade.objects.all())
        self.cidades = {norm(c.nome): c for c in _cids}
        # Fallback sem espaços: o export da planilha às vezes gruda os espaços
        # ("RioDoCampo" → "Rio do Campo"). Só mapeia para cidade REAL da base —
        # é normalização de espaço, não invenção de dado. Ambíguo → não entra.
        _ns = {}
        for c in _cids:
            k = norm(c.nome).replace(' ', '')
            _ns[k] = None if k in _ns else c
        self.cidades_nospace = {k: v for k, v in _ns.items() if v is not None}
        self.users = {}
        for u in Usuario.objects.filter(is_active=True):
            self.users[norm(u.username)] = u
        self.stats['usuarios_no_sistema'] = len(self.users)

        est_rows, est_key = self._rows('BASE DA ESTRUTURA', wb, TITULOS_ESTRUTURA)
        ver_rows, ver_key = self._rows('BASE DE CONTATOS VERDADEIRA', wb, TITULOS_VERDADEIRA)

        # Nomes que só existem na VERDADEIRA → leads pendentes
        est_nomes = {norm(cell_str(self._v(r, est_key, 'nome')))
                     for r in est_rows if cell_str(self._v(r, est_key, 'nome'))}
        exclusivos = [r for r in ver_rows
                      if cell_str(self._v(r, ver_key, 'nome'))
                      and norm(cell_str(self._v(r, ver_key, 'nome'))) not in est_nomes]
        self.stats['exclusivos_verdadeira'] = len(exclusivos)

        modo = 'GRAVANDO' if executar else 'DRY-RUN (nada será gravado)'
        self.stdout.write(self.style.WARNING(f'== {modo} =='))

        with transaction.atomic():
            for r in est_rows:
                self._importar_estrutura(r, est_key, executar)
            for r in exclusivos:
                self._importar_lead_verdadeira(r, ver_key, executar)
            if not executar:
                transaction.set_rollback(True)

        self.stdout.write('')
        self.stdout.write(self.style.MIGRATE_HEADING('== Resultado =='))
        for k in sorted(self.stats):
            self.stdout.write(f'  {k}: {self.stats[k]}')
        if self.atendentes_sem:
            self.stdout.write(self.style.WARNING(
                f'\n== Atendentes SEM usuário no sistema ({len(self.atendentes_sem)} nomes distintos) =='))
            for nome, n in self.atendentes_sem.most_common():
                self.stdout.write(f'  "{nome}": {n} linha(s)')
        if self.cidades_sem:
            self.stdout.write(self.style.WARNING(
                f'\n== Cidades SEM correspondência na base ({len(self.cidades_sem)} distintas) =='))
            for cid, n in self.cidades_sem.most_common(40):
                self.stdout.write(f'  {cid}: {n} linha(s)')
            if len(self.cidades_sem) > 40:
                self.stdout.write(f'  … e mais {len(self.cidades_sem) - 40} cidades')
        self.stdout.write(self.style.SUCCESS('\nGravado.') if executar
                          else self.style.WARNING('\nDRY-RUN — rode com --sim para gravar.'))

    # ── helpers de leitura ──
    def _rows(self, aba, wb, titulos):
        rows = list(wb[aba].iter_rows(values_only=True))
        hdr = rows[2]  # cabeçalho na linha 3
        key = {}
        for i, t in enumerate(hdr):
            n = norm(t)
            if n in titulos:
                key[titulos[n]] = i
        return rows[3:], key

    def _v(self, row, key, chave):
        i = key.get(chave)
        return row[i] if (i is not None and i < len(row)) else None

    def _atendente(self, row, key):
        nome = cell_str(self._v(row, key, 'atendente'))
        if not nome:
            self.stats['atendente_vazio'] += 1
            return None, ''
        u = self.users.get(norm(nome))
        if u:
            self.stats['atendente_vinculado'] += 1
            return u, ''
        self.stats['atendente_sem_usuario'] += 1
        self.atendentes_sem[nome] += 1
        return None, f'atendente "{nome}" sem usuário'

    def _cidade(self, row, key):
        raw = cell_str(self._v(row, key, 'cidade'))
        uf = cell_str(self._v(row, key, 'uf')).upper()
        c = self.cidades.get(norm(raw)) if raw else None
        if raw and not c:
            c = self.cidades_nospace.get(norm(raw).replace(' ', ''))
            if c:
                self.stats['cidade_recuperada_sem_espaco'] += 1
        if raw and not c:
            self.cidades_sem[f'{raw}/{uf or "?"}'] += 1
        return c, raw, uf

    # ── ESTRUTURA → aprovado ──
    def _importar_estrutura(self, row, key, executar):
        nome = cell_str(self._v(row, key, 'nome'))
        if not nome:
            return
        self.stats['estrutura_linhas'] += 1
        cidade, cid_raw, uf = self._cidade(row, key)
        atd, nota_atd = self._atendente(row, key)
        obs = cell_str(self._v(row, key, 'observacoes'))
        avisos = []
        if not cidade and cid_raw:
            avisos.append(f'Cidade fora da base: {cid_raw}/{uf or "?"}')
            self.stats['sem_cidade'] += 1
        if nota_atd:
            avisos.append(nota_atd)
        if avisos:
            obs = (obs + '\n' if obs else '') + '⚠️ ' + ' · '.join(avisos)

        defaults = self._defaults_comuns(row, key, uf, atd, obs)
        hist = cell_str(self._v(row, key, 'historico'))
        obj = self._upsert(nome, cidade, uf, defaults, aprovacao='aprovado', executar=executar)
        self.stats['estrutura_' + ('criados' if obj[1] else 'atualizados')] += 1
        if hist and obj[0] is not None:
            self._historico(obj[0], hist, defaults.get('data_contato'), executar)

    # ── VERDADEIRA exclusiva → lead pendente ──
    def _importar_lead_verdadeira(self, row, key, executar):
        nome = cell_str(self._v(row, key, 'nome'))
        self.stats['lead_linhas'] += 1
        cidade, cid_raw, uf = self._cidade(row, key)
        atd, nota_atd = self._atendente(row, key)
        obs = cell_str(self._v(row, key, 'observacoes'))
        # Diferenças da VERDADEIRA → observação de revisão
        difs = []
        rede2 = cell_str(self._v(row, key, 'rede2'))
        engaj = cell_str(self._v(row, key, 'engajamento'))
        if rede2:
            difs.append(f'Rede social 2: {rede2}')
        if engaj:
            difs.append(f'Nível engajamento (planilha): {engaj}')
        if not cidade and cid_raw:
            difs.append(f'Cidade fora da base: {cid_raw}/{uf or "?"}')
        if nota_atd:
            difs.append(nota_atd)
        aviso = '⚠️ REVISAR — lead da BASE DE CONTATOS VERDADEIRA.'
        if difs:
            aviso += ' ' + ' · '.join(difs)
        obs = (obs + '\n' if obs else '') + aviso

        defaults = self._defaults_comuns(row, key, uf, atd, obs)
        obj = self._upsert(nome, cidade, uf, defaults, aprovacao='pendente', executar=executar)
        self.stats['lead_' + ('criados' if obj[1] else 'atualizados')] += 1

    def _defaults_comuns(self, row, key, uf, atd, obs):
        return {
            'intencao_voto': map_voto(self._v(row, key, 'voto')),
            'nivel': map_nivel(self._v(row, key, 'nivel')),   # ausente na VERDADEIRA → ''
            'uf': (uf or 'SC')[:2],
            'contato_feito': to_bool(self._v(row, key, 'contato_feito')),
            'data_contato': to_date(self._v(row, key, 'data_contato')),
            'canal_atendimento': map_canal(self._v(row, key, 'canal')),
            'vaquinha_enviada': to_bool(self._v(row, key, 'vaquinha')),
            'doou': to_bool(self._v(row, key, 'doou')),
            'observacoes': obs,
            'filiado_partido': cell_str(self._v(row, key, 'filiado'))[:100],
            'quem_e_eleitor': cell_str(self._v(row, key, 'quem_eleitor'))[:200],
            'origem_contato': cell_str(self._v(row, key, 'origem'))[:200],
            'instagram': cell_str(self._v(row, key, 'instagram'))[:100],
            'facebook': cell_str(self._v(row, key, 'facebook'))[:100],
            'email': clean_email(self._v(row, key, 'email')),
            'endereco': cell_str(self._v(row, key, 'endereco'))[:300],
            'material_entregue': to_bool(self._v(row, key, 'material')),
            'idade': to_idade(self._v(row, key, 'idade')),
            'segmentos': cell_str(self._v(row, key, 'segmentos'))[:255],
            'telefone': clean_phone(self._v(row, key, 'telefone')),
            'atendente_user': atd,
        }

    def _upsert(self, nome, cidade, uf, defaults, aprovacao, executar):
        telefone = defaults.get('telefone', '')
        qs = Lideranca.all_objects.filter(papel='apoiador')
        exist = qs.filter(telefone=telefone).first() if telefone else None
        if not exist:
            f = {'nome__iexact': nome}
            f['cidade'] = cidade if cidade else None
            exist = qs.filter(**f).first()
        if exist:
            if executar:
                for c, v in defaults.items():
                    if v not in (None, '', False) or c in ('contato_feito', 'doou',
                                                           'vaquinha_enviada', 'material_entregue'):
                        setattr(exist, c, v)
                exist.save()
            return exist, False
        obj = Lideranca(papel='apoiador', nome=nome, cidade=cidade,
                        regiao=cidade.regiao if cidade else None,
                        tipo='comunitario', status='ativo', origem='import',
                        aprovacao=aprovacao, **defaults)
        if executar:
            obj.save()
            return obj, True
        return None, True

    def _historico(self, obj, texto, quando, executar):
        if not executar:
            return
        data = timezone.now()
        if quando:
            data = timezone.make_aware(datetime.combine(quando, datetime.min.time()))
        if InteracaoLog.objects.filter(lideranca=obj, descricao=texto).exists():
            return
        InteracaoLog.objects.create(lideranca=obj, tipo='outro', descricao=texto, data=data)
        self.stats['historico'] += 1
